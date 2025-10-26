import openpyxl
workbook = openpyxl.load_workbook("../buoi 3/logindata.xlsx")
sheet = workbook.active
# valueC2 = sheet['D2'].value
# print(valueC2)
for row in sheet.iter_rows(min_row=2, values_only=True):
    stt, student_id, name, score = row
    #print(f"stt: {stt}")
    #Lấy dữ liệu của cột
column_score = [sheet.cell(row= r, column= 4 ).value for r in range(1, sheet.max_row+1)]
print(column_score)

row_four = [sheet.cell(row = 4, column= c).value for c in range(1, sheet.max_column+1)]
print(row_four)
